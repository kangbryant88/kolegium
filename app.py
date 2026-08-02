from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, make_response, jsonify
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import OperationalError
from flask_migrate import Migrate
from datetime import datetime, date, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
import io
import os
import pandas as pd
import openpyxl
from docx import Document
from fpdf import FPDF
from flask_mail import Mail, Message
import secrets
import urllib.parse
from dotenv import load_dotenv

from config import Config
from models import db, Rol, Usuario, Anuncio, Bitacora, PlanificacionDefensoria, Grado, Tema, AsistenciaDiaria, AsistenciaPersonal, Representante, Estudiante, EnlaceTemporal, Incidencia, AsistenciaEstudiante, AlertaDefensoria, Brigada, Acta, SolicitudEnlace, SolicitudActualizacion, SolicitudDefensoria, TokenRecuperacion, ConfiguracionInstitucional

app = Flask(__name__)
app.config.from_object(Config)

# Aseguramos que la carpeta de uploads exista (obtenido de Config)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Inicializamos las extensiones con la app
from extensions import mail
mail.init_app(app)
db.init_app(app)
migrate = Migrate(app, db)

# ==========================================
# --- ESTRATEGIA DE BLUEPRINTS (PRÓXIMA FASE) ---
# ==========================================
# Para desmantelar el resto de app.py, propondremos crear:
# 1. auth_bp (Login, Registro, Espera, Logout)
# 2. admin_bp (Gestión de usuarios y roles)
# 3. academico_bp (Estadísticas globales, planificador docente, asistencia)
# 4. defensoria_bp (Incidencias, bitácoras, actas)
# ==========================================

# ==========================================
# --- 2. INICIALIZACIÓN DE LA BD ---
# ==========================================

with app.app_context():
    db.create_all()  # Inicialización Automática: Crea las tablas faltantes si el archivo db es nuevo
    # Mapa maestro de permisos por rol
    PERMISOS_POR_ROL = {
        'Administrador Supremo': 'dashboard,planificador,asistencia,defensoria,configuracion,admin,anuncios',
        'Equipo Directivo': 'dashboard,asistencia,configuracion,admin,anuncios',
        'Administrativo': 'dashboard_general,asistencia,anuncios,planificador',
        'Docente de Aula': 'planificador,asistencia,anuncios',
        'Docente Especialista': 'planificador,asistencia,anuncios',
        'Defensoría Estudiantil': 'dashboard_general,defensoria',
        'Obrero': 'dashboard_general',
        'Personal de Vigilancia': 'dashboard_general',
        'Personal de Cocina': 'dashboard_general',
    }

    if not Rol.query.first():
        for nombre, permisos in PERMISOS_POR_ROL.items():
            db.session.add(Rol(nombre=nombre, permisos=permisos))
        db.session.commit()
    else:
        # Migración: sincronizar permisos de roles existentes y crear faltantes
        for nombre, permisos in PERMISOS_POR_ROL.items():
            rol = Rol.query.filter_by(nombre=nombre).first()
            if rol:
                rol.permisos = permisos
            else:
                db.session.add(Rol(nombre=nombre, permisos=permisos))
                
        # Limpiar rol duplicado "Vigilante" si existe
        rol_vig = Rol.query.filter_by(nombre='Vigilante').first()
        rol_pv = Rol.query.filter_by(nombre='Personal de Vigilancia').first()
        if rol_vig and rol_pv:
            for u in Usuario.query.filter_by(rol_id=rol_vig.id).all():
                u.rol_id = rol_pv.id
            db.session.delete(rol_vig)
            
        # Limpiar rol obsoleto "Coordinador / Administrativo" si existe
        rol_coord = Rol.query.filter_by(nombre='Coordinador / Administrativo').first()
        rol_admin = Rol.query.filter_by(nombre='Administrativo').first()
        if rol_coord and rol_admin:
            for u in Usuario.query.filter_by(rol_id=rol_coord.id).all():
                u.rol_id = rol_admin.id
            db.session.delete(rol_coord)
            
        db.session.commit()

@app.context_processor
def inyectar_datos_globales():
    return dict(
        mis_permisos=session.get('permisos', ''),
        mi_rol=session.get('nombre_rol', 'Sin Rol'),
        total_anuncios=Anuncio.query.count(),
        hoy_str=datetime.now().strftime('%d/%m/%Y')
    )

def obtener_estudiantes_por_docente(usuario_id):
    grado = Grado.query.filter(Grado.docentes.any(id=usuario_id)).first()
    if grado:
        return Estudiante.query.filter_by(grado_id=grado.id).order_by(Estudiante.nombre_completo.asc()).all()
    return []

def auth_defensoria():
    if not session.get('logeado'): return False
    
    nombre_rol = session.get('nombre_rol')
    depto = session.get('departamento_asignado')
    
    if nombre_rol in ['Defensoría Estudiantil', 'Equipo Directivo', 'Administrador Supremo']:
        return True
        
    if nombre_rol == 'Administrativo' and depto == 'Defensoría':
        return True
        
    return False

# ==========================================
# --- 3. DASHBOARD Y AUTENTICACIÓN ---
# ==========================================

@app.route('/')
def index():
    if not session.get('logeado'): return render_template('landing.html')
    permisos = session.get('permisos', '')

    ahora = datetime.now()
    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    fecha_hoy_esp = f"{dias[ahora.weekday()]}, {ahora.day} de {meses[ahora.month-1]} de {ahora.year}"

    registros_hoy = AsistenciaDiaria.query.filter_by(fecha=ahora.date()).all()
    mat_hoy = sum(r.matricula_total for r in registros_hoy)
    asist_hoy = sum(r.asistentes for r in registros_hoy)
    porc_hoy = round((asist_hoy / mat_hoy) * 100, 1) if mat_hoy > 0 else 0.0

    total_v = sum(g.total_varones for g in Grado.query.all())
    total_h = sum(g.total_hembras for g in Grado.query.all())
    ultimos_anuncios = Anuncio.query.order_by(Anuncio.fecha.desc()).limit(3).all()

    if 'dashboard' not in permisos.split(','):
        config_inst = ConfiguracionInstitucional.query.first()
        return render_template('dashboard_general.html', fecha_full=fecha_hoy_esp,
                               matricula_hoy=mat_hoy, asistentes_hoy=asist_hoy, 
                               porcentaje_hoy=porc_hoy, anuncios=ultimos_anuncios,
                               total_v=total_v, total_h=total_h, config_inst=config_inst)

    registros_pers_hoy = AsistenciaPersonal.query.filter_by(fecha=date.today()).all()
    mat_pers_hoy = sum(r.matricula_base for r in registros_pers_hoy)
    asist_pers_hoy = sum(r.asistentes for r in registros_pers_hoy)
    porc_pers_hoy = round((asist_pers_hoy / mat_pers_hoy) * 100, 1) if mat_pers_hoy > 0 else 0.0
    
    desglose_personal = {'Docentes': 0, 'Administrativos': 0, 'Obreros': 0, 'Cocina': 0, 'Vigilantes': 0}
    for r in registros_pers_hoy:
        if r.categoria in desglose_personal: desglose_personal[r.categoria] += r.asistentes

    total_v = sum(g.total_varones for g in Grado.query.all())
    total_h = sum(g.total_hembras for g in Grado.query.all())
    ultimos_anuncios = Anuncio.query.order_by(Anuncio.fecha.desc()).limit(3).all()

    historico = AsistenciaDiaria.query.order_by(AsistenciaDiaria.fecha.desc()).limit(5).all()
    historico.reverse()
    labels_g = [h.fecha.strftime('%d/%m') for h in historico]
    datos_g = [h.porcentaje for h in historico]

    solicitudes_pendientes = SolicitudEnlace.query.filter_by(estado='Pendiente').all()
    actualizaciones_pendientes = SolicitudActualizacion.query.filter_by(estado='Pendiente').all()
    solicitudes_defensoria = SolicitudDefensoria.query.filter_by(estado='Pendiente').order_by(SolicitudDefensoria.fecha_solicitud.asc()).all()
    actividades_recientes = Bitacora.query.order_by(Bitacora.fecha.desc()).limit(4).all()
    return render_template('index.html', fecha_full=fecha_hoy_esp, matricula_hoy=mat_hoy, 
                           asistentes_hoy=asist_hoy, porcentaje_hoy=porc_hoy,
                           total_v=total_v, total_h=total_h, anuncios=ultimos_anuncios,
                           labels_g=labels_g, datos_g=datos_g, mat_pers_hoy=mat_pers_hoy,
                           asist_pers_hoy=asist_pers_hoy, porc_pers_hoy=porc_pers_hoy,
                           desglose_personal=desglose_personal, 
                           solicitudes_pendientes=solicitudes_pendientes,
                           actualizaciones_pendientes=actualizaciones_pendientes, 
                           solicitudes_defensoria=solicitudes_defensoria,
                           actividades_recientes=actividades_recientes)

@app.route('/modulo_construccion')
def modulo_construccion():
    if not session.get('logeado'): return redirect(url_for('auth.login'))
    return render_template('construccion.html')

@app.route('/dashboard_general')
def ver_dashboard_general():
    if not session.get('logeado'): return redirect(url_for('auth.login'))
    
    ahora = datetime.now()
    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    fecha_hoy_esp = f"{dias[ahora.weekday()]}, {ahora.day} de {meses[ahora.month-1]} de {ahora.year}"

    registros_hoy = AsistenciaDiaria.query.filter_by(fecha=ahora.date()).all()
    mat_hoy = sum(r.matricula_total for r in registros_hoy)
    asist_hoy = sum(r.asistentes for r in registros_hoy)
    porc_hoy = round((asist_hoy / mat_hoy) * 100, 1) if mat_hoy > 0 else 0.0

    total_v = sum(g.total_varones for g in Grado.query.all())
    total_h = sum(g.total_hembras for g in Grado.query.all())
    ultimos_anuncios = Anuncio.query.order_by(Anuncio.fecha.desc()).limit(3).all()

    config_inst = ConfiguracionInstitucional.query.first()
    
    return render_template('dashboard_general.html', fecha_full=fecha_hoy_esp,
                           matricula_hoy=mat_hoy, asistentes_hoy=asist_hoy, 
                           porcentaje_hoy=porc_hoy, anuncios=ultimos_anuncios,
                           total_v=total_v, total_h=total_h, config_inst=config_inst)

@app.route('/dashboard/aprobar_solicitud/<int:id>', methods=['POST'])
def aprobar_solicitud(id):
    solicitud = SolicitudDefensoria.query.get_or_404(id)
    solicitud.estado = 'Aprobada'
    solicitud.fecha_respuesta = datetime.utcnow()
    db.session.commit()
    flash('Solicitud de acceso aprobada.', 'success')
    return redirect(url_for('index'))

@app.route('/dashboard/rechazar_solicitud/<int:id>', methods=['POST'])
def rechazar_solicitud(id):
    solicitud = SolicitudDefensoria.query.get_or_404(id)
    solicitud.estado = 'Rechazada'
    solicitud.fecha_respuesta = datetime.utcnow()
    db.session.commit()
    flash('Solicitud rechazada.', 'danger')
    return redirect(url_for('index'))

# ==========================================
# --- REGISTRO DE BLUEPRINTS ---
# ==========================================
from routes.auth import auth_bp
app.register_blueprint(auth_bp)

from routes.admin import admin_bp
app.register_blueprint(admin_bp)

from routes.academico import academico_bp
app.register_blueprint(academico_bp)

# ==========================================
# --- 2. INICIALIZACIÓN DE LA BD ---
# ==========================================

# ==========================================
# --- 6. DEFENSORÍA ESTUDIANTIL ---
# ==========================================

@app.route('/defensoria', methods=['GET', 'POST'])
def defensoria():
    if not auth_defensoria():
        flash("Acceso denegado: No tienes permisos para ingresar a Defensoría.")
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        db.session.add(PlanificacionDefensoria(
            tema_charla=request.form['tema_charla'], proposito=request.form['proposito'],
            poblacion_objetivo=request.form['poblacion_objetivo'], usuario_id=session['usuario_id'],
            fecha_programada=datetime.strptime(request.form['fecha_programada'], '%Y-%m-%d').date(), 
            hora_programada=request.form['hora_programada']
        ))
        db.session.commit(); return redirect(url_for('defensoria'))
        
    planes = PlanificacionDefensoria.query.filter_by(usuario_id=session['usuario_id']).all()
    estudiantes = Estudiante.query.order_by(Estudiante.nombre_completo).all()
    brigadas = Brigada.query.all()
    actas = Acta.query.order_by(Acta.fecha.desc()).all()
    solicitudes = SolicitudDefensoria.query.order_by(SolicitudDefensoria.fecha_solicitud.desc()).all()
    grados = Grado.query.order_by(Grado.nombre).all()
    alertas = AlertaDefensoria.query.order_by(AlertaDefensoria.fecha_emision.desc()).all()
    return render_template('defensoria.html', planificaciones=planes, estudiantes=estudiantes, brigadas=brigadas, actas=actas, solicitudes=solicitudes, grados=grados, alertas=alertas)

@app.route('/editar_defensoria/<int:id>', methods=['POST'])
def editar_defensoria(id):
    reg = PlanificacionDefensoria.query.get_or_404(id)
    reg.tema_charla, reg.proposito, reg.poblacion_objetivo = request.form['tema_charla'], request.form['proposito'], request.form['poblacion_objetivo']
    db.session.commit(); return redirect(url_for('defensoria'))

@app.route('/eliminar_defensoria/<int:id>', methods=['POST'])
def eliminar_defensoria(id):
    db.session.delete(PlanificacionDefensoria.query.get_or_404(id)); db.session.commit()
    return redirect(url_for('defensoria'))

@app.route('/defensoria/solicitar_datos', methods=['POST'])
def solicitar_datos():
    if not auth_defensoria(): return "Acceso Denegado", 403
    estudiante_id = request.form.get('estudiante_id')
    motivo = request.form.get('motivo')
    solicitante_id = session.get('usuario_id')
    
    if estudiante_id and motivo:
        nueva_solicitud = SolicitudDefensoria(
            estudiante_id=estudiante_id,
            solicitante_id=solicitante_id,
            motivo=motivo
        )
        db.session.add(nueva_solicitud)
        db.session.commit()
        flash("Solicitud enviada a Dirección exitosamente.", "success")
        
    return redirect(url_for('defensoria'))

@app.route('/defensoria/eliminar_solicitud/<int:id>', methods=['POST'])
def eliminar_solicitud_def(id):
    if not auth_defensoria(): return "Acceso Denegado", 403
    solicitud = SolicitudDefensoria.query.get_or_404(id)
    db.session.delete(solicitud)
    db.session.commit()
    flash("Solicitud eliminada correctamente.", "success")
    return redirect(url_for('defensoria'))

@app.route('/defensoria/alertar_docente/<int:id>', methods=['POST'])
def alertar_docente_incidencia(id):
    if not auth_defensoria(): return "Acceso Denegado", 403
    solicitud = SolicitudDefensoria.query.get_or_404(id)
    nueva_alerta = AlertaDefensoria(
        estudiante_id=solicitud.estudiante_id,
        motivo=f"Defensoría Incidencia: {solicitud.motivo}",
        estatus_atencion='Pendiente'
    )
    db.session.add(nueva_alerta)
    db.session.commit()
    flash("Alerta enviada al docente exitosamente.", "success")
    return redirect(url_for('defensoria'))

@app.route('/defensoria/descargar_ficha_pdf/<int:id>')
def descargar_ficha_pdf(id):
    if not auth_defensoria(): return "Acceso Denegado", 403
    solicitud = SolicitudDefensoria.query.get_or_404(id)
    if solicitud.estado != 'Aprobada':
        flash('Esta solicitud no ha sido aprobada por la Dirección.', 'danger')
        return redirect(url_for('defensoria'))
    current_date = datetime.now().strftime('%d/%m/%Y %I:%M %p')
    return render_template('pdf_ficha_visita.html', solicitud=solicitud, current_date=current_date)

@app.route('/guardar_acta', methods=['POST'])
def guardar_acta():
    if not auth_defensoria(): return "Acceso Denegado", 403
    contenido = request.form.get('contenido_manual')
    tipo = request.form.get('tipo_acta', 'General')
    est_id = request.form.get('estudiante_id')
    estudiante_id = int(est_id) if est_id else None
    
    nueva_acta = Acta(
        contenido_manual=contenido,
        tipo_acta=tipo,
        estudiante_id=estudiante_id,
        defensor_id=session['usuario_id']
    )
    db.session.add(nueva_acta)
    db.session.commit()
    return redirect(url_for('defensoria'))

@app.route('/eliminar_acta/<int:id>', methods=['POST'])
def eliminar_acta(id):
    if not auth_defensoria(): return "Acceso Denegado", 403
    db.session.delete(Acta.query.get_or_404(id)); db.session.commit()
    return redirect(url_for('defensoria'))

@app.route('/crear_brigada', methods=['POST'])
def crear_brigada():
    if not auth_defensoria(): return "Acceso Denegado", 403
    nombre = request.form.get('nombre')
    desc = request.form.get('descripcion')
    if nombre:
        db.session.add(Brigada(nombre=nombre, descripcion=desc))
        db.session.commit()
    return redirect(url_for('defensoria'))

@app.route('/asignar_brigada', methods=['POST'])
def asignar_brigada():
    if not auth_defensoria(): return "Acceso Denegado", 403
    estudiante_id = request.form.get('estudiante_id')
    brigada_id = request.form.get('brigada_id')
    
    if estudiante_id and brigada_id:
        estudiante = Estudiante.query.get(estudiante_id)
        brigada = Brigada.query.get(brigada_id)
        if estudiante and brigada and estudiante not in brigada.estudiantes:
            brigada.estudiantes.append(estudiante)
            db.session.commit()
    return redirect(url_for('defensoria'))

# --- EXPORTAR PDF DEFENSORÍA ---
@app.route('/exportar_pdf_defensoria')
def exportar_pdf_defensoria():
    if not session.get('logeado'): return redirect(url_for('auth.login'))
    
    # Creamos el documento PDF
    pdf = FPDF()
    pdf.add_page()
    
    # Título del Documento
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, txt="Cronograma de Defensoria Estudiantil", ln=True, align='C')
    pdf.set_font("Arial", 'I', 12)
    pdf.cell(200, 10, txt=f"Docente: {session.get('nombre_completo')}", ln=True, align='C')
    pdf.ln(10)
    
    # Encabezados de la tabla
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(25, 10, 'Fecha', 1, 0, 'C')
    pdf.cell(25, 10, 'Hora', 1, 0, 'C')
    pdf.cell(40, 10, 'Grado/Seccion', 1, 0, 'C')
    pdf.cell(50, 10, 'Tema de Charla', 1, 0, 'C')
    pdf.cell(50, 10, 'Proposito', 1, 1, 'C')
    
    # Consultamos los datos del usuario activo
    planes = PlanificacionDefensoria.query.filter_by(usuario_id=session['usuario_id']).order_by(PlanificacionDefensoria.fecha_programada.asc()).all()
    
    # Llenamos la tabla
    pdf.set_font("Arial", '', 9)
    for p in planes:
        # Acortamos los textos si son muy largos para que no rompan la tabla
        tema = (p.tema_charla[:25] + '...') if len(p.tema_charla) > 25 else p.tema_charla
        proposito = (p.proposito[:25] + '...') if len(p.proposito) > 25 else p.proposito
        
        pdf.cell(25, 10, p.fecha_programada.strftime('%d/%m/%Y'), 1)
        pdf.cell(25, 10, p.hora_programada, 1)
        pdf.cell(40, 10, p.poblacion_objetivo, 1)
        pdf.cell(50, 10, tema, 1)
        pdf.cell(50, 10, proposito, 1)
        pdf.ln()
        
    # Preparamos la descarga
    response = make_response(pdf.output(dest='S').encode('latin-1'))
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=Cronograma_Defensoria.pdf'
    return response    

# --- CARGA MASIVA EXCEL ---
@app.route('/descargar_plantilla')
def descargar_plantilla():
    output = io.BytesIO(); wb = openpyxl.Workbook(); ws = wb.active
    ws.append(['TEMA_CHARLA', 'PROPOSITO_OBJETIVO']); wb.save(output); output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                     as_attachment=True, download_name='Plantilla_Defensoria.xlsx')

@app.route('/subir_plan', methods=['POST'])
def subir_plan():
    archivo = request.files.get('archivo_plan')
    if not archivo: return redirect(url_for('defensoria'))
    wb = openpyxl.load_workbook(archivo); ws = wb.active
    extraidas = [{'tema': str(r[0]).strip(), 'proposito': str(r[1]).strip()} for r in ws.iter_rows(min_row=2, values_only=True) if r[0] and r[1]]
    return render_template('asignar_masivo.html', charlas=extraidas)

@app.route('/guardar_masivo', methods=['POST'])
def guardar_masivo():
    temas, propositos, grados, fechas, horas = request.form.getlist('tema[]'), request.form.getlist('proposito[]'), request.form.getlist('grado[]'), request.form.getlist('fecha[]'), request.form.getlist('hora[]')
    for i in range(len(temas)):
        if grados[i] and fechas[i] and horas[i]:
            db.session.add(PlanificacionDefensoria(tema_charla=temas[i], proposito=propositos[i], poblacion_objetivo=grados[i], 
                                                  fecha_programada=datetime.strptime(fechas[i], '%Y-%m-%d').date(), 
                                                  hora_programada=horas[i], usuario_id=session['usuario_id']))
    db.session.commit(); return redirect(url_for('defensoria'))

@app.route('/configuracion', methods=['GET', 'POST'])
def configuracion():
    if not session.get('logeado'): return redirect(url_for('auth.login'))
    if request.method == 'POST':
        if 'grado_num' in request.form and 'seccion_letra' in request.form:
            docente_ids = request.form.getlist('docente_ids[]')
            nombre_completo = f"{request.form.get('grado_num')} {request.form.get('seccion_letra')}"
            nuevo_grado = Grado(nombre=nombre_completo, 
                                 total_varones=int(request.form.get('m_varones') or 0),
                                 total_hembras=int(request.form.get('m_hembras') or 0))
            if docente_ids:
                docentes_asignados = Usuario.query.filter(Usuario.id.in_(docente_ids)).all()
                nuevo_grado.docentes.extend(docentes_asignados)
            db.session.add(nuevo_grado)
        if 'nuevo_tema' in request.form:
            db.session.add(Tema(nombre=request.form['nuevo_tema'], usuario_id=session['usuario_id']))
        if 'config_inst' in request.form:
            conf = ConfiguracionInstitucional.query.first()
            if not conf:
                conf = ConfiguracionInstitucional()
                db.session.add(conf)
            conf.nombre_escuela = request.form.get('nombre_escuela')
            conf.director = request.form.get('director')
            conf.telefono_director = request.form.get('telefono_director')
            conf.correo_director = request.form.get('correo_director')
            conf.codigo_estadistico = request.form.get('codigo_estadistico')
            conf.codigo_dea = request.form.get('codigo_dea')
            conf.codigo_administrativo = request.form.get('codigo_administrativo')
            conf.codigo_dependencia = request.form.get('codigo_dependencia')
            conf.codigo_sunagro = request.form.get('codigo_sunagro')
            conf.rif_escuela = request.form.get('rif_escuela')
            conf.rif_consejo = request.form.get('rif_consejo')
            conf.dependencia = request.form.get('dependencia')
            conf.ubicacion_geografica = request.form.get('ubicacion_geografica')
            conf.clase_plantel = request.form.get('clase_plantel')
            conf.ano_fundacion = request.form.get('ano_fundacion')
            conf.telefono_escuela = request.form.get('telefono_escuela')
            conf.correo_escuela = request.form.get('correo_escuela')
            conf.supervisora = request.form.get('supervisora')
            conf.direccion = request.form.get('direccion')
            conf.circuito = request.form.get('circuito')

        db.session.commit(); return redirect(url_for('configuracion'))
        
    docentes = Usuario.query.join(Rol).filter(Rol.nombre == 'Docente de Aula').all()
    config_inst = ConfiguracionInstitucional.query.first()
    return render_template('configuracion.html', grados=Grado.query.all(), temas=Tema.query.all(), docentes=docentes, config_inst=config_inst)

@app.route('/editar_grado/<int:id>', methods=['POST'])
def editar_grado(id):
    g = Grado.query.get_or_404(id)
    if 'grado_num' in request.form and 'seccion_letra' in request.form:
        g.nombre = f"{request.form.get('grado_num')} {request.form.get('seccion_letra')}"
    g.total_varones, g.total_hembras = int(request.form.get('m_varones', 0)), int(request.form.get('m_hembras', 0))
    docente_ids = request.form.getlist('docente_ids[]')
    if docente_ids:
        g.docentes = Usuario.query.filter(Usuario.id.in_(docente_ids)).all()
    else:
        g.docentes = []
    db.session.commit(); return redirect(url_for('configuracion'))

@app.route('/borrar_grado/<int:id>')
def borrar_grado(id):
    db.session.delete(Grado.query.get_or_404(id)); db.session.commit()
    return redirect(url_for('configuracion'))

@app.route('/borrar_tema/<int:id>')
def borrar_tema(id):
    db.session.delete(Tema.query.get_or_404(id)); db.session.commit()
    return redirect(url_for('configuracion'))

@app.route('/configuracion/promover_ano', methods=['POST'])
def promover_ano():
    if not session.get('logeado') or session.get('nombre_rol') not in ['Administrador Supremo', 'Equipo Directivo']:
        return redirect(url_for('index'))
        
    estudiantes_activos = Estudiante.query.filter_by(estatus='Activo').all()
    grados_existentes = {g.nombre: g.id for g in Grado.query.all()}
    
    # Pre-validación
    movimientos_requeridos = set()
    mapeo_grados = {
        '1er Grado': '2do Grado',
        '2do Grado': '3er Grado',
        '3er Grado': '4to Grado',
        '4to Grado': '5to Grado',
        '5to Grado': '6to Grado'
    }
    
    for est in estudiantes_activos:
        if not est.grado_id: continue
        grado_actual = Grado.query.get(est.grado_id)
        if not grado_actual: continue
        
        if '6to Grado' in grado_actual.nombre:
            continue
            
        for nivel_actual, nivel_siguiente in mapeo_grados.items():
            if nivel_actual in grado_actual.nombre:
                nombre_nuevo_grado = grado_actual.nombre.replace(nivel_actual, nivel_siguiente)
                movimientos_requeridos.add(nombre_nuevo_grado)
                break
                
    grados_faltantes = [grado for grado in movimientos_requeridos if grado not in grados_existentes]
    
    if grados_faltantes:
        flash(f"Operación cancelada: Faltan crear los siguientes grados de destino: {', '.join(grados_faltantes)}. Por favor créelos primero.", "danger")
        return redirect(url_for('configuracion'))
        
    # Ejecución Atómica
    for est in estudiantes_activos:
        if not est.grado_id: continue
        grado_actual = Grado.query.get(est.grado_id)
        if not grado_actual: continue
        
        if '6to Grado' in grado_actual.nombre:
            est.estatus = 'Egreso'
            if est.procedencia:
                est.procedencia += f" | Cohorte Egresada {date.today().year}-{date.today().year+1}"
            else:
                est.procedencia = f"Cohorte Egresada {date.today().year}-{date.today().year+1}"
        else:
            for nivel_actual, nivel_siguiente in mapeo_grados.items():
                if nivel_actual in grado_actual.nombre:
                    nombre_nuevo_grado = grado_actual.nombre.replace(nivel_actual, nivel_siguiente)
                    est.grado_id = grados_existentes[nombre_nuevo_grado]
                    break
                    
    try:
        db.session.commit()
        flash("Promoción masiva ejecutada con éxito. Todos los estudiantes han sido promovidos o egresados.", "success")
    except Exception as e:
        db.session.rollback()
        flash("Ocurrió un error inesperado durante la promoción. Se han revertido los cambios.", "danger")
        
    return redirect(url_for('configuracion'))

# ==========================================
# --- 9. ANUNCIOS ---
# ==========================================

@app.route('/anuncios', methods=['GET', 'POST'])
def anuncios():
    if not session.get('logeado'): return redirect(url_for('auth.login'))
    if request.method == 'POST' and 'anuncios' in session.get('permisos', ''):
        db.session.add(Anuncio(titulo=request.form['titulo'], mensaje=request.form['mensaje'], autor_id=session['usuario_id']))
        db.session.commit(); return redirect(url_for('anuncios'))
    return render_template('anuncios.html', anuncios=Anuncio.query.order_by(Anuncio.fecha.desc()).all())

@app.route('/eliminar_anuncio/<int:id>', methods=['POST'])
def eliminar_anuncio(id):
    if 'admin' in session.get('permisos', ''):
        db.session.delete(Anuncio.query.get(id)); db.session.commit()
    return redirect(url_for('anuncios'))

# ==========================================
# --- 11. GESTIÓN DE PERSONAL ---
# ==========================================

@app.route('/gestion_personal')
def gestion_personal():
    if not session.get('logeado') or 'admin' not in session.get('permisos', []):
        return redirect(url_for('auth.login'))
        
    personal = Usuario.query.all()
    
    # Calculate HR Document Compliance Metrics
    base_metric = lambda: {'docentes': 0, 'operativos': 0, 'directivos': 0, 'faltantes': 0}
    metrics = {
        'rif': base_metric(),
        'cv': base_metric(),
        'constancia': base_metric(),
        'voucher': base_metric(),
        'cedula': base_metric()
    }
    
    macro_docentes = ['Docente de Aula (1ro a 6to)', 'Especialista (Robótica / Deportes)', 'Defensoría Estudiantil']
    macro_directivos = ['Equipo Directivo / Administrativo', 'Administrador Supremo']
    macro_operativos = ['Obrero', 'Personal de Vigilancia', 'Personal de Cocina']
    
    for t in personal:
        area = t.area_trabajo or ''
        
        if area in macro_directivos:
            cat = 'directivos'
        elif area in macro_docentes:
            cat = 'docentes'
        elif area in macro_operativos:
            cat = 'operativos'
        else:
            cat = 'faltantes' # Fallback for unassigned or broken roles
            
        # If cat is faltantes, we still just attribute it to faltantes for charts, 
        # but to avoid breaking the dict, let's say they just add to missing unless they have it?
        # Actually if they don't belong to a category, they are outside the macros.
        # But we must count their docs. Let's just put them in 'docentes' as a safe default or ignore their loaded state?
        # Better: if cat is not identified, we just skip or put them in faltantes
        if cat != 'faltantes':
            if t.rif_path: metrics['rif'][cat] += 1
            else: metrics['rif']['faltantes'] += 1
            
            if t.curriculum_path: metrics['cv'][cat] += 1
            else: metrics['cv']['faltantes'] += 1
            
            if t.constancia_path: metrics['constancia'][cat] += 1
            else: metrics['constancia']['faltantes'] += 1
            
            if t.voucher_path: metrics['voucher'][cat] += 1
            else: metrics['voucher']['faltantes'] += 1
            
            if t.cedula_path: metrics['cedula'][cat] += 1
            else: metrics['cedula']['faltantes'] += 1
        else:
            # If they don't have a valid role, everything counts as faltante
            metrics['rif']['faltantes'] += 1
            metrics['cv']['faltantes'] += 1
            metrics['constancia']['faltantes'] += 1
            metrics['voucher']['faltantes'] += 1
            metrics['cedula']['faltantes'] += 1
            
    return render_template('gestion_personal.html', personal=personal, metrics=metrics)

@app.route('/actualizar_credenciales_mppe/<int:id>', methods=['POST'])
def actualizar_credenciales_mppe(id):
    if not session.get('logeado') or 'admin' not in session.get('permisos', []):
        return redirect(url_for('auth.login'))
        
    usuario = Usuario.query.get_or_404(id)
    correo_mppe = request.form.get('correo_mppe')
    clave_mppe = request.form.get('clave_mppe')
    
    usuario.usuario_autogestion = correo_mppe
    usuario.clave_autogestion = clave_mppe
    
    db.session.commit()
    flash('Credenciales del MPPE actualizadas correctamente.', 'success')
    return redirect(url_for('gestion_personal'))

from werkzeug.utils import secure_filename

# ==========================================
# --- 12. PORTAL DEL TRABAJADOR ---
# ==========================================

@app.route('/portal_trabajador', methods=['GET', 'POST'])
def portal_trabajador():
    if not session.get('logeado'):
        return redirect(url_for('auth.login'))
        
    usuario = Usuario.query.get(session['usuario_id'])
    
    if request.method == 'POST':
        tipo_documento = request.form.get('tipo_documento')
        if 'archivo' not in request.files:
            flash('No se seleccionó ningún archivo.', 'error')
            return redirect(url_for('portal_trabajador'))
            
        file = request.files['archivo']
        if file.filename == '':
            flash('No se seleccionó ningún archivo.', 'error')
            return redirect(url_for('portal_trabajador'))
            
        extension = file.filename.split('.')[-1].lower()
        if file and extension in ['pdf', 'png', 'jpg', 'jpeg']:
            # Patrón de nombre estandarizado: [tipo_documento]_[id_usuario].[extension]
            prefix = f"{tipo_documento}_{usuario.id}"
            
            # Buscar y eliminar cualquier archivo previo con el mismo prefijo
            for existing_file in os.listdir(app.config['UPLOAD_FOLDER']):
                if existing_file.startswith(prefix + "."):
                    old_path = os.path.join(app.config['UPLOAD_FOLDER'], existing_file)
                    if os.path.exists(old_path): 
                        os.remove(old_path)
            
            # Asignar nuevo nombre
            safe_filename = f"{prefix}.{extension}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
            file.save(file_path)
            
            if tipo_documento == 'voucher':
                usuario.voucher_path = safe_filename
            elif tipo_documento == 'constancia':
                usuario.constancia_path = safe_filename
            elif tipo_documento == 'curriculum':
                usuario.curriculum_path = safe_filename
            elif tipo_documento == 'rif':
                usuario.rif_path = safe_filename
            elif tipo_documento == 'cedula':
                usuario.cedula_path = safe_filename
            elif tipo_documento == 'foto_perfil':
                usuario.foto_perfil_path = safe_filename
                
            db.session.commit()
            flash('Documento guardado exitosamente en tu expediente.', 'success')
            return redirect(url_for('portal_trabajador'))
        else:
            flash('Tipo de archivo no permitido. Solo PDF, JPG o PNG.', 'error')
            return redirect(url_for('portal_trabajador'))

    return render_template('portal_trabajador.html', usuario=usuario)

from flask import send_from_directory

@app.route('/ver_documento_personal/<filename>')
def ver_documento_personal(filename):
    if not session.get('logeado'):
        return redirect(url_for('auth.login'))
    # Verify permissions: must be admin or the owner of the document
    if 'admin' not in session.get('permisos', '') and not filename.startswith(f"{session.get('usuario_id')}_"):
        flash("Acceso denegado al documento.", "error")
        return redirect(url_for('index'))
        
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/mi_perfil', methods=['GET', 'POST'])
def mi_perfil():
    if not session.get('logeado'):
        return redirect(url_for('auth.login'))
        
    usuario = Usuario.query.get(session['usuario_id'])
    
    if request.method == 'POST':
        nombre_completo = request.form.get('nombre_completo')
        area_trabajo = request.form.get('area_trabajo')
        cedula = request.form.get('cedula')
        file = request.files.get('foto_perfil')
        
        if nombre_completo:
            usuario.nombre_completo = nombre_completo
            session['nombre_completo'] = nombre_completo
            
        if cedula:
            usuario.cedula = cedula
            
        if area_trabajo and area_trabajo != usuario.area_trabajo and not usuario.cargo_solicitado:
            usuario.cargo_solicitado = area_trabajo
            
        if file and file.filename.split('.')[-1].lower() in ['png', 'jpg', 'jpeg']:
            import os
            extension = file.filename.split('.')[-1].lower()
            prefix = f"foto_perfil_{usuario.id}"
            # Replace existing file
            for existing_file in os.listdir(app.config['UPLOAD_FOLDER']):
                if existing_file.startswith(prefix + "."):
                    old_path = os.path.join(app.config['UPLOAD_FOLDER'], existing_file)
                    if os.path.exists(old_path): 
                        os.remove(old_path)
            
            safe_filename = f"{prefix}.{extension}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
            file.save(file_path)
            usuario.foto_perfil_path = safe_filename
            session['foto_perfil_path'] = safe_filename
            
        db.session.commit()
        flash('Perfil actualizado exitosamente.', 'success')
        return redirect(url_for('mi_perfil'))
        
    return render_template('mi_perfil.html', usuario=usuario)

if __name__ == '__main__':
    app.run(debug=True)